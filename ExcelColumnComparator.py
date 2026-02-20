#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from pathlib import Path
import os
import subprocess
import sys

from utils.theme import COLORS as C, F, F_BOLD, F_TITLE, F_SUB, F_SMALL, F_STAT, apply_styles
from utils.text import normalize
from utils.excel import load_dataframe, resolve_sheet_name, get_filter_header_rows
from utils.comparison import collect_col_data, get_rows_with_unique_values, row_matches_search
from utils.treeview import auto_size_columns
from utils.template import show_template_validation_card


class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Column Comparator")
        self.root.geometry("1280x820")
        self.root.configure(bg=C["bg"])
        self.root.option_add("*TCombobox*Listbox.background", C["input_bg"])
        self.root.option_add("*TCombobox*Listbox.foreground", C["text"])
        self.root.option_add("*TCombobox*Listbox.selectBackground", C["accent"])
        self.root.option_add("*TCombobox*Listbox.selectForeground", "#fff")
        self.root.withdraw()
        apply_styles()
        self.file_configs = []
        self.history = []
        self.temp_files = []
        self.template_mode = False
        self.template_files = []
        self.template_config = None
        self.pick_files()
        self.root.mainloop()

    def clear(self):
        for w in self.root.winfo_children():
            w.destroy()

    # ── File picker ──────────────────────────────────────────────────────
    def pick_files(self):
        self.root.withdraw()
        paths = filedialog.askopenfilenames(
            title="Select Excel files",
            filetypes=[("Excel", "*.xlsx *.xls")])
        if not paths:
            if len(self.file_configs) >= 2:
                self.show_pair_selector(); return
            self.root.destroy(); return
        self.temp_files = list(paths)
        if len(paths) > 1 and not self.template_mode:
            if messagebox.askyesno("", "Do you want to create a template?"):
                self.template_mode = True
                self.template_files = list(paths)
        self.process_next_file()

    def process_next_file(self):
        if not self.temp_files:
            if len(self.file_configs) >= 2:
                self.show_pair_selector()
            else:
                self.root.destroy()
            return
        self.cur_path = self.temp_files.pop(0)
        self.wb = load_workbook(self.cur_path, data_only=True)
        self.show_sheet_and_header()

    # ── Sheet + header row ───────────────────────────────────────────────
    def show_sheet_and_header(self):
        self.root.deiconify()
        self.clear()
        self.history.append("sheet_and_header")

        bar = tk.Frame(self.root, bg=C["bar"], height=80)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text=Path(self.cur_path).name, font=F_TITLE,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=24, pady=16)
        tk.Label(bar, text=f"File {len(self.file_configs)+1}  |  Step 1 : Choose a sheet and click the header row",
                 font=F_SUB, fg=C["dim"], bg=C["bar"]).pack(side=tk.RIGHT, padx=24)

        mid = tk.Frame(self.root, bg=C["bg"])
        mid.pack(fill=tk.BOTH, expand=True, padx=16, pady=12)

        sb = tk.Frame(mid, bg=C["surface"], width=200,
                      highlightbackground=C["border"], highlightthickness=1)
        sb.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10)); sb.pack_propagate(False)
        tk.Label(sb, text="Sheets", font=F_BOLD, bg=C["surface"],
                 fg=C["cyan"]).pack(padx=12, pady=(12, 6), anchor="w")
        self.sheet_lb = tk.Listbox(sb, font=F, relief="flat", bd=0,
                                   selectbackground=C["accent"], selectforeground="#fff",
                                   highlightthickness=0, bg=C["surface"], fg=C["text"])
        for name in self.wb.sheetnames:
            suffix = "  [hidden]" if self.wb[name].sheet_state == "hidden" else ""
            self.sheet_lb.insert(tk.END, f"  {name}{suffix}")
        self.sheet_lb.select_set(0)
        self.sheet_lb.bind("<<ListboxSelect>>", lambda _: self.load_sheet())
        self.sheet_lb.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        ct = tk.Frame(mid, bg=C["surface"],
                      highlightbackground=C["border"], highlightthickness=1)
        ct.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        xscr = ttk.Scrollbar(ct, orient=tk.HORIZONTAL)
        yscr = ttk.Scrollbar(ct, orient=tk.VERTICAL)
        self.tree = ttk.Treeview(ct, style="T.Treeview", show="headings",
                                 xscrollcommand=xscr.set, yscrollcommand=yscr.set)
        xscr.config(command=self.tree.xview); yscr.config(command=self.tree.yview)
        yscr.pack(side=tk.RIGHT, fill=tk.Y)
        xscr.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<<TreeviewSelect>>", self.on_row_select)

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=16, pady=(0, 12))
        ttk.Button(bot, text="  Back  ", style="A.TButton",
                   command=self.go_back).pack(side=tk.LEFT)
        self.status = tk.StringVar(value="Click a row to mark it as the table header")
        tk.Label(bot, textvariable=self.status, font=F, bg=C["bg"],
                 fg=C["dim"]).pack(side=tk.LEFT, expand=True, padx=16)
        self.next_btn = ttk.Button(bot, text="  Next  ", style="A.TButton",
                                   command=self.confirm_header, state=tk.DISABLED)
        self.next_btn.pack(side=tk.RIGHT)
        self.sel_row = None
        self.load_sheet()

    def selected_sheet_name(self):
        idx = self.sheet_lb.curselection()
        return self.wb.sheetnames[idx[0]] if idx else self.wb.sheetnames[0]

    def load_sheet(self):
        ws = self.wb[self.selected_sheet_name()]
        self.filter_rows = get_filter_header_rows(ws)
        self.raw = [tuple(cell.value for cell in row)
                    for row in ws.iter_rows(max_row=200, max_col=40)]
        if not self.raw:
            return
        num_cols = max(len(r) for r in self.raw)
        col_ids = [f"c{i}" for i in range(num_cols)]
        headings = [f"Col {i+1}" for i in range(num_cols)]
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = col_ids
        display_values = []
        for i, row in enumerate(self.raw):
            vals = [normalize(x) if x is not None else "" for x in row]
            vals += [""] * (num_cols - len(vals))
            display_values.append(vals)
            tag = "filter_row" if i in self.filter_rows else ("alt" if i % 2 else "normal")
            self.tree.insert("", "end", iid=str(i), values=vals, tags=(tag,))
        for i, cid in enumerate(col_ids):
            self.tree.heading(cid, text=headings[i])
        auto_size_columns(self.tree, col_ids, display_values, headings)
        self.tree.tag_configure("filter_row", background=C["orange"], foreground=C["bar"])
        self.tree.tag_configure("alt", background=C["alt"])
        self.tree.tag_configure("normal", background=C["surface"])
        self.sel_row = None
        self.status.set("Click a row to mark it as the table header")
        self.next_btn.config(state=tk.DISABLED)

    def on_row_select(self, _):
        sel = self.tree.selection()
        if not sel:
            return
        self.sel_row = int(sel[0])
        for i, iid in enumerate(self.tree.get_children()):
            tag = "filter_row" if i in self.filter_rows else ("alt" if i % 2 else "normal")
            self.tree.item(iid, tags=(tag,))
        self.tree.tag_configure("filter_row", background=C["orange"], foreground=C["bar"])
        self.tree.tag_configure("alt", background=C["alt"], foreground=C["text"])
        self.tree.tag_configure("normal", background=C["surface"], foreground=C["text"])
        self.tree.item(sel[0], tags=("sel",))
        self.tree.tag_configure("sel", background=C["sel"], foreground="#fff")
        preview = [normalize(v) for v in self.raw[self.sel_row] if v is not None][:6]
        self.status.set(f"Row {self.sel_row+1}:  {' | '.join(preview)}")
        self.next_btn.config(state=tk.NORMAL)

    def confirm_header(self):
        self.cur_sheet = self.selected_sheet_name()
        self.cur_hdr_idx = self.sel_row
        self.cur_hdr_vals = self.raw[self.cur_hdr_idx]
        self.wb.close()
        self.show_column_selection()

    # ── Column selection ─────────────────────────────────────────────────
    def show_column_selection(self):
        self.root.deiconify()
        self.clear()
        self.history.append("col_sel")

        bar = tk.Frame(self.root, bg=C["bar"], height=80)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text=Path(self.cur_path).name, font=F_TITLE,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=24, pady=16)
        tk.Label(bar, text=f"File {len(self.file_configs)+1}  |  Step 2 : Select columns to compare",
                 font=F_SUB, fg=C["dim"], bg=C["bar"]).pack(side=tk.RIGHT, padx=24)

        card = tk.Frame(self.root, bg=C["surface"],
                        highlightbackground=C["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=20, pady=16)
        tk.Label(card, text="Tick one or more columns:", font=F_BOLD,
                 bg=C["surface"], fg=C["cyan"]).pack(padx=20, pady=(20, 10), anchor="w")

        wrap = tk.Frame(card, bg=C["surface"])
        wrap.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        self.col_vars = []
        col_index = 0
        cols_per_row = 4
        for val in self.cur_hdr_vals:
            if val is None:
                continue
            name = normalize(val)
            if not name or name.lower() == "nan":
                continue
            var = tk.BooleanVar()
            cb = tk.Checkbutton(wrap, text=name, variable=var, font=F,
                                bg=C["surface"], fg=C["text"],
                                selectcolor=C["chk_sel"],
                                activebackground=C["surface"],
                                activeforeground=C["accent"],
                                padx=10, pady=6, anchor="w")
            r, c = divmod(col_index, cols_per_row)
            cb.grid(row=r, column=c, sticky="w", padx=6, pady=3)
            self.col_vars.append((name, var))
            col_index += 1
        for c in range(cols_per_row):
            wrap.columnconfigure(c, weight=1)

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=20, pady=(0, 16))
        ttk.Button(bot, text="  Back  ", style="A.TButton",
                   command=self.go_back).pack(side=tk.LEFT)
        ttk.Button(bot, text="  Next  ", style="A.TButton",
                   command=self.confirm_columns).pack(side=tk.RIGHT)

    def confirm_columns(self):
        cols = [n for n, v in self.col_vars if v.get()]
        if not cols:
            messagebox.showwarning("", "Select at least one column"); return
        self.file_configs.append({"path": self.cur_path, "sheet": self.cur_sheet,
                                  "header_row": self.cur_hdr_idx, "columns": cols})
        if self.template_mode and not self.template_config:
            self.template_config = {"sheet": self.cur_sheet,
                                    "header_row": self.cur_hdr_idx, "columns": cols}
            for tfile in self.template_files[1:]:
                self.file_configs.append({"path": tfile, "sheet": self.cur_sheet,
                                         "header_row": self.cur_hdr_idx, "columns": cols})
            self.generate_template_mappings()
            self.run_comparison()
        elif self.temp_files:
            self.process_next_file()
        elif messagebox.askyesno("", f"{len(self.file_configs)} file(s) added.\nAdd another file?"):
            self.pick_files()
        elif len(self.file_configs) >= 2:
            self.show_pair_selector()
        else:
            messagebox.showinfo("", "Need at least 2 files"); self.pick_files()

    def generate_template_mappings(self):
        cols = self.template_config["columns"]
        self.mappings = {i: [(c, c) for c in cols]
                         for i in range(1, len(self.file_configs))}

    # ── Pair selector ────────────────────────────────────────────────────
    def show_pair_selector(self):
        self.root.deiconify()
        self.clear()
        self.history.append("pair_selector")
        SKIP = "-- skip --"

        bar = tk.Frame(self.root, bg=C["bar"], height=80)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text="What compares to what?", font=F_TITLE,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=24, pady=16)
        tk.Label(bar, text="Final step : Map columns from File #1 to each other file",
                 font=F_SUB, fg=C["dim"], bg=C["bar"]).pack(side=tk.RIGHT, padx=24)

        card = tk.Frame(self.root, bg=C["surface"],
                        highlightbackground=C["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=20, pady=12)

        f1 = self.file_configs[0]
        others = self.file_configs[1:]

        canvas = tk.Canvas(card, bg=C["surface"], highlightthickness=0)
        vsb = ttk.Scrollbar(card, orient=tk.VERTICAL, command=canvas.yview)
        inner = tk.Frame(canvas, bg=C["surface"])
        inner.bind("<Configure>", lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw", tags="inner")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure("inner", width=e.width))
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 4), pady=4)
        canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.bind_scroll(canvas)

        grid = tk.Frame(inner, bg=C["surface"])
        grid.pack(fill=tk.X, padx=8, pady=8)

        f1_name = Path(f1["path"]).name
        tk.Label(grid, text=f"#1  {f1_name}", font=F_BOLD,
                 bg=C["surface"], fg=C["accent"]).grid(row=0, column=0, sticky="w",
                                                       padx=(10, 20), pady=(0, 8))
        for k, oth in enumerate(others):
            tk.Label(grid, text=f"#{k+2}  {Path(oth['path']).name}", font=F_BOLD,
                     bg=C["surface"], fg=C["orange"]).grid(row=0, column=k+1,
                                                           sticky="w", padx=10, pady=(0, 8))

        tk.Frame(grid, bg=C["border"], height=1).grid(
            row=1, column=0, columnspan=1+len(others), sticky="ew", padx=10, pady=(0, 6))

        self.map_vars = []
        for ri, col1 in enumerate(f1["columns"]):
            tk.Label(grid, text=col1, font=F, bg=C["surface"],
                     fg=C["text"]).grid(row=ri+2, column=0, sticky="w", padx=(10, 20), pady=4)
            row_vars = []
            for k, oth in enumerate(others):
                options = [SKIP] + oth["columns"]
                var = tk.StringVar(value=SKIP)
                for oc in oth["columns"]:
                    if oc.strip().lower() == col1.strip().lower():
                        var.set(oc); break
                ttk.Combobox(grid, textvariable=var, values=options,
                             state="readonly", width=28, font=F).grid(
                    row=ri+2, column=k+1, sticky="w", padx=10, pady=4)
                row_vars.append((k+1, var))
            self.map_vars.append((col1, row_vars))

        for c in range(1+len(others)):
            grid.columnconfigure(c, weight=1)

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=20, pady=(0, 16))
        ttk.Button(bot, text="  Back  ", style="A.TButton",
                   command=self.go_back).pack(side=tk.LEFT)
        ttk.Button(bot, text="  Compare  ", style="A.TButton",
                   command=self.confirm_pairs).pack(side=tk.RIGHT)

    def confirm_pairs(self):
        SKIP = "-- skip --"
        self.mappings = {}
        for col1, row_vars in self.map_vars:
            for fidx, var in row_vars:
                if var.get() != SKIP:
                    self.mappings.setdefault(fidx, []).append((col1, var.get()))
        if not self.mappings:
            messagebox.showwarning("", "Map at least one column"); return
        self.run_comparison()

    # ── Comparison results ───────────────────────────────────────────────
    def run_comparison(self):
        self.root.deiconify()
        self.clear()

        bar = tk.Frame(self.root, bg=C["bar"], height=80)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text="Comparison Results", font=F_TITLE,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=24, pady=16)

        dfs = [load_dataframe(cfg["path"], cfg["header_row"],
                              resolve_sheet_name(cfg["path"], cfg["sheet"]))
               for cfg in self.file_configs]

        canvas = tk.Canvas(self.root, bg=C["bg"], highlightthickness=0)
        vsb = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=canvas.yview)
        results_frame = tk.Frame(canvas, bg=C["bg"])
        results_frame.bind("<Configure>",
                           lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=results_frame, anchor="nw", tags="inner")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure("inner", width=e.width))
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        self.bind_scroll(canvas)

        f1_cfg = self.file_configs[0]
        f1_name = Path(f1_cfg["path"]).name
        summary_data = []

        for fidx, col_pairs in sorted(self.mappings.items()):
            fN_cfg = self.file_configs[fidx]
            fN_name = Path(fN_cfg["path"]).name

            col_data_1 = {c1: collect_col_data(dfs[0], c1, f1_cfg["header_row"])
                          for c1, _ in col_pairs}
            col_data_N = {cN: collect_col_data(dfs[fidx], cN, fN_cfg["header_row"])
                          for _, cN in col_pairs}

            all_vals_1 = {v for data in col_data_1.values() for v in data.values()}
            all_vals_N = {v for data in col_data_N.values() for v in data.values()}
            common   = all_vals_1 & all_vals_N
            unique_1 = all_vals_1 - all_vals_N
            unique_N = all_vals_N - all_vals_1

            rows_only_1 = get_rows_with_unique_values(col_data_1, unique_1)
            rows_only_N = get_rows_with_unique_values(col_data_N, unique_N)

            card = tk.Frame(results_frame, bg=C["surface"],
                            highlightbackground=C["border"], highlightthickness=1)
            card.pack(fill=tk.X, pady=8, padx=6)

            tk.Label(card,
                     text=f"{f1_name} [{f1_cfg['sheet']}]   vs   {fN_name} [{fN_cfg['sheet']}]",
                     font=F_BOLD, bg=C["surface"], fg=C["text"]).pack(padx=20, pady=(16, 4), anchor="w")

            mapped_str = "    ".join(f"{c1}  \u2192  {cN}" for c1, cN in col_pairs)
            tk.Label(card, text=mapped_str, font=F,
                     bg=C["surface"], fg=C["dim"]).pack(padx=20, pady=(0, 8), anchor="w")

            if self.template_mode:
                show_template_validation_card(card, self.template_config, fN_cfg, fN_name)

            stats = tk.Frame(card, bg=C["surface"])
            stats.pack(fill=tk.X, padx=20, pady=6)
            self.make_stat_badge(stats, "Common", len(common), C["green"])
            self.make_stat_badge(stats, f"Only in {f1_name}", len(rows_only_1), C["accent"])
            self.make_stat_badge(stats, f"Only in {fN_name}", len(rows_only_N), C["orange"])

            cols_1 = [c1 for c1, _ in col_pairs]
            cols_N = [cN for _, cN in col_pairs]

            if rows_only_1:
                self.build_result_grid(card, f"Only in {f1_name}", cols_1,
                                       rows_only_1, C["accent"], f1_cfg["path"])
            if rows_only_N:
                self.build_result_grid(card, f"Only in {fN_name}", cols_N,
                                       rows_only_N, C["orange"], fN_cfg["path"])

            tk.Frame(card, bg=C["surface"], height=12).pack()
            summary_data.append((f1_name, fN_name, len(common),
                                 len(rows_only_1), len(rows_only_N), fidx+1))

        self.build_summary(results_frame, summary_data)

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=16, pady=12)
        ttk.Button(bot, text="  New Comparison  ", style="A.TButton",
                   command=self.new_comparison).pack(side=tk.RIGHT)

    def make_stat_badge(self, parent, label, count, color):
        badge = tk.Frame(parent, bg=C["badge_bg"],
                         highlightbackground=color, highlightthickness=1)
        badge.pack(side=tk.LEFT, padx=(0, 12))
        tk.Label(badge, text=f"  {label}:  {count}  ", font=F_STAT,
                 bg=C["badge_bg"], fg=color).pack(padx=4, pady=4)

    def build_summary(self, parent, summary_data):
        sm = tk.Frame(parent, bg=C["hdr_bg"],
                      highlightbackground=C["border"], highlightthickness=1)
        sm.pack(fill=tk.X, pady=8, padx=6)
        tk.Label(sm, text="Summary", font=F_TITLE, fg=C["purple"],
                 bg=C["hdr_bg"]).pack(padx=20, pady=(16, 8), anchor="w")
        for f1_n, fN_n, common_c, u1_c, uN_c, num in summary_data:
            tk.Label(sm,
                     text=f"{f1_n} vs {fN_n}:  {common_c} common,  "
                          f"{u1_c} unique to #1,  {uN_c} unique to #{num}",
                     font=F, fg=C["dim"], bg=C["hdr_bg"]).pack(padx=20, pady=2, anchor="w")
        tk.Frame(sm, bg=C["hdr_bg"], height=16).pack()

    def build_result_grid(self, parent, title, col_names, rows_data, color, file_path):
        fr = tk.Frame(parent, bg=C["surface"])
        fr.pack(fill=tk.X, padx=20, pady=(8, 4))

        hdr = tk.Frame(fr, bg=C["surface"])
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=f"{title}:", font=F_BOLD, bg=C["surface"],
                 fg=color).pack(side=tk.LEFT, anchor="w")
        tk.Button(hdr, text="\u2197 Open in Excel", font=F_SMALL,
                  bg=C["surface2"], fg=color, relief="flat", cursor="hand2",
                  activebackground=C["alt"], activeforeground=color, bd=0,
                  command=lambda: self.open_file(file_path)).pack(side=tk.LEFT, padx=12)

        search_frame = tk.Frame(fr, bg=C["surface"])
        search_frame.pack(fill=tk.X, pady=(6, 4))
        tk.Label(search_frame, text="Search:", font=F,
                 bg=C["surface"], fg=C["dim"]).pack(side=tk.LEFT, padx=(0, 4))
        search_var = tk.StringVar()
        tk.Entry(search_frame, textvariable=search_var, font=F,
                 bg=C["input_bg"], fg=C["text"], relief="flat", bd=0,
                 insertbackground=C["accent"], highlightbackground=C["input_bd"],
                 highlightcolor=C["accent"], highlightthickness=1).pack(
            side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True, ipady=4)
        tk.Label(search_frame, text="A11*=starts  *A11*=contains  *A11=ends  A11=exact",
                 font=F_SMALL, bg=C["surface"], fg=C["dim"]).pack(side=tk.LEFT)

        grid_frame = tk.Frame(fr, bg=C["surface"])
        grid_frame.pack(fill=tk.X, pady=(4, 4))

        all_columns = ["Row"] + col_names
        tree = ttk.Treeview(grid_frame, columns=all_columns, show="headings",
                            style="T.Treeview", height=min(len(rows_data), 10))
        yscr = ttk.Scrollbar(grid_frame, orient=tk.VERTICAL, command=tree.yview)
        xscr = ttk.Scrollbar(grid_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscr.set, xscrollcommand=xscr.set)

        for col in all_columns:
            tree.heading(col, text=col)
        tree.column("Row", anchor="center", stretch=False)

        all_rows = [[str(row)] + [rows_data[row].get(col, "") for col in col_names]
                    for row in sorted(rows_data)]

        for i, values in enumerate(all_rows):
            tree.insert("", "end", values=values, tags=("alt" if i % 2 else "normal",))

        auto_size_columns(tree, all_columns, all_rows, all_columns)
        tree.tag_configure("alt", background=C["alt"])
        tree.tag_configure("normal", background=C["surface"])

        search_var.trace_add("write", lambda *_: self.filter_tree(tree, all_rows, search_var))
        tree.bind("<Double-1>", lambda e: self.on_cell_double_click(tree, e))

        yscr.pack(side=tk.RIGHT, fill=tk.Y)
        xscr.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(fill=tk.X)

    def filter_tree(self, tree, all_rows, search_var):
        pattern = search_var.get().strip()
        tree.delete(*tree.get_children())
        for i, values in enumerate(all_rows):
            if not pattern or row_matches_search(values, pattern):
                tree.insert("", "end", values=values, tags=("alt" if i % 2 else "normal",))

    def on_cell_double_click(self, tree, event):
        item = tree.identify_row(event.y)
        col = tree.identify_column(event.x)
        if not item or not col:
            return
        col_index = int(col.replace("#", "")) - 1
        values = tree.item(item, "values")
        if values and col_index < len(values) and values[col_index]:
            self.copy_to_clipboard(values[col_index])

    def copy_to_clipboard(self, text):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        original_title = self.root.title()
        self.root.title(f"Copied: {text}")
        self.root.after(1500, lambda: self.root.title(original_title))

    def open_file(self, file_path):
        if sys.platform == "win32":
            os.startfile(file_path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])

    def bind_scroll(self, canvas):
        def scroll_widget(event, delta):
            w = self.root.winfo_containing(event.x_root, event.y_root)
            while w is not None:
                if isinstance(w, ttk.Treeview):
                    w.yview_scroll(delta, "units")
                    return "break"
                w = getattr(w, "master", None)
            canvas.yview_scroll(delta, "units")
        canvas.bind_all("<MouseWheel>", lambda e: scroll_widget(e, -e.delta // 120))
        canvas.bind_all("<Button-4>", lambda e: scroll_widget(e, -3))
        canvas.bind_all("<Button-5>", lambda e: scroll_widget(e, 3))

    def new_comparison(self):
        self.file_configs = []
        self.history = []
        self.temp_files = []
        self.template_mode = False
        self.template_files = []
        self.template_config = None
        self.pick_files()

    def go_back(self):
        if len(self.history) < 2:
            return
        self.history.pop()
        prev = self.history[-1]
        if prev == "sheet_and_header":
            self.show_sheet_and_header()
        elif prev == "col_sel":
            self.show_column_selection()
        elif prev == "pair_selector":
            self.show_pair_selector()


if __name__ == "__main__":
    App()
