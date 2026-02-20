#!/usr/bin/env python3
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import os
import subprocess
import sys

# ── Light Theme ──────────────────────────────────────────────────────────
C = {
    "bg":      "#f0f2f5", "surface":  "#ffffff", "surface2": "#f8f9fb",
    "bar":     "#2c3e50", "bar_text": "#ecf0f1", "accent":   "#3498db",
    "accent_dk": "#2980b9", "text": "#2c3e50", "dim": "#7f8c8d",
    "border":  "#dce1e6", "sel": "#3498db", "alt": "#f8f9fa",
    "hdr_bg":  "#34495e", "green": "#27ae60", "orange": "#e67e22",
    "red": "#e74c3c", "cyan": "#2980b9", "purple": "#8e44ad",
    "chk_sel": "#ffffff",
}
F      = ("Segoe UI", 10)
F_B    = ("Segoe UI", 10, "bold")
F_T    = ("Segoe UI", 16, "bold")
F_SUB  = ("Segoe UI", 12)


class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Column Comparator")
        self.root.geometry("1200x780")
        self.root.configure(bg=C["bg"])
        self.root.withdraw()
        self._styles()
        self.file_configs = []
        self.history = []
        self.temp_files = []
        self.template_mode = False
        self.template_files = []
        self.template_config = None
        self.pick_files()
        self.root.mainloop()

    def _styles(self):
        s = ttk.Style(); s.theme_use("clam")
        s.configure("T.Treeview", background=C["surface"], foreground=C["text"],
                     fieldbackground=C["surface"], font=F, rowheight=26)
        s.configure("T.Treeview.Heading", background=C["hdr_bg"],
                     foreground=C["cyan"], font=F_B, relief="flat")
        s.map("T.Treeview.Heading", background=[("active", C["surface2"])])
        s.map("T.Treeview", background=[("selected", C["sel"])],
              foreground=[("selected", "#fff")])
        s.configure("A.TButton", background=C["accent"], foreground="#fff",
                     font=F_B, padding=(20, 10))
        s.map("A.TButton", background=[("active", C["accent_dk"])])
        s.configure("TScrollbar", background=C["surface2"], troughcolor=C["bg"],
                     bordercolor=C["bg"], arrowcolor=C["dim"])

    def clear(self):
        for w in self.root.winfo_children():
            w.destroy()

    # ── 1  File picker ───────────────────────────────────────────────────
    def pick_files(self):
        self.root.withdraw()
        paths = filedialog.askopenfilenames(
            title="Select Excel files",
            filetypes=[("Excel", "*.xlsx *.xls")])
        if not paths:
            if len(self.file_configs) >= 2:
                self.show_pair_selector(); return
            self.root.destroy(); return
        self.temp_files = list(paths)
        if len(paths) > 1 and not self.template_mode:
            if messagebox.askyesno("", "Do you want to create a template?"):
                self.template_mode = True
                self.template_files = list(paths)
        self.process_next_file()

    def process_next_file(self):
        if not self.temp_files:
            if len(self.file_configs) >= 2:
                self.show_pair_selector()
            else:
                self.root.destroy()
            return
        self.cur_path = self.temp_files.pop(0)
        self.wb = load_workbook(self.cur_path, data_only=True)
        self.show_sheet_and_header()

    # ── 2  Sheet + header row ────────────────────────────────────────────
    def show_sheet_and_header(self):
        self.root.deiconify()
        self.clear()
        self.history.append("sheet_and_header")
        bar = tk.Frame(self.root, bg=C["bar"], height=75)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text=Path(self.cur_path).name, font=F_T,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=20)
        tk.Label(bar, text=f"File {len(self.file_configs)+1}  |  Step 1 : Choose a sheet and click the header row",
                 font=F_SUB, fg=C["dim"], bg=C["bar"]).pack(side=tk.RIGHT, padx=20)

        mid = tk.Frame(self.root, bg=C["bg"])
        mid.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        # sheet list
        sb = tk.Frame(mid, bg=C["surface"], width=190,
                      highlightbackground=C["border"], highlightthickness=1)
        sb.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 8)); sb.pack_propagate(False)
        tk.Label(sb, text="Sheets", font=F_B, bg=C["surface"],
                 fg=C["cyan"]).pack(padx=10, pady=(8, 4), anchor="w")
        self.sheet_lb = tk.Listbox(sb, font=F, relief="flat", bd=0,
                                   selectbackground=C["accent"], selectforeground="#fff",
                                   highlightthickness=0, bg=C["surface"], fg=C["text"])
        for name in self.wb.sheetnames:
            lock = "  [hidden]" if self.wb[name].sheet_state == "hidden" else ""
            self.sheet_lb.insert(tk.END, f"  {name}{lock}")
        self.sheet_lb.select_set(0)
        self.sheet_lb.bind("<<ListboxSelect>>", lambda _: self.load_sheet())
        self.sheet_lb.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # treeview
        ct = tk.Frame(mid, bg=C["surface"],
                      highlightbackground=C["border"], highlightthickness=1)
        ct.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        xscr = ttk.Scrollbar(ct, orient=tk.HORIZONTAL)
        yscr = ttk.Scrollbar(ct, orient=tk.VERTICAL)
        self.tree = ttk.Treeview(ct, style="T.Treeview", show="headings",
                                 xscrollcommand=xscr.set, yscrollcommand=yscr.set)
        xscr.config(command=self.tree.xview); yscr.config(command=self.tree.yview)
        yscr.pack(side=tk.RIGHT, fill=tk.Y)
        xscr.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(fill=tk.BOTH, expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_row)

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=12, pady=(0, 8))
        back_btn = ttk.Button(bot, text="<- Back", style="A.TButton", command=self._back)
        back_btn.pack(side=tk.LEFT)
        self.status = tk.StringVar(value="Click a row to mark it as the table header")
        tk.Label(bot, textvariable=self.status, font=F, bg=C["bg"],
                 fg=C["dim"]).pack(side=tk.LEFT, expand=True)
        self.next_btn = ttk.Button(bot, text="Next  ->", style="A.TButton",
                                   command=self._header_next, state=tk.DISABLED)
        self.next_btn.pack(side=tk.RIGHT)
        self.sel_row = None
        self.load_sheet()

    def _sname(self):
        idx = self.sheet_lb.curselection()
        return self.wb.sheetnames[idx[0]] if idx else self.wb.sheetnames[0]

    def load_sheet(self):
        ws = self.wb[self._sname()]
        self.raw, self.rcols = [], []
        for row in ws.iter_rows(max_row=200, max_col=40):
            vals, clr = [], None
            for cell in row:
                vals.append(cell.value)
                if clr is None and cell.fill and cell.fill.start_color:
                    try:
                        rgb = cell.fill.start_color.rgb
                        if isinstance(rgb, str) and len(rgb) == 8 and rgb != "00000000":
                            clr = f"#{rgb[2:]}"
                    except Exception:
                        pass
            self.raw.append(tuple(vals)); self.rcols.append(clr)
        if not self.raw:
            return
        nc = max(len(r) for r in self.raw)
        cids = [f"c{i}" for i in range(nc)]
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = cids
        for i, c in enumerate(cids):
            self.tree.heading(c, text=f"Col {i+1}")
            self.tree.column(c, width=120, minwidth=50)
        for i, row in enumerate(self.raw):
            v = [str(x) if x is not None else "" for x in row] + [""] * (nc - len(row))
            tag = f"r{i}"
            self.tree.insert("", "end", iid=str(i), values=v, tags=(tag,))
            bg = self.rcols[i] or (C["alt"] if i % 2 else C["surface"])
            self.tree.tag_configure(tag, background=bg)
        self.sel_row = None
        self.status.set("Click a row to mark it as the table header")
        self.next_btn.config(state=tk.DISABLED)

    def _on_row(self, _):
        sel = self.tree.selection()
        if not sel:
            return
        self.sel_row = int(sel[0])
        for i, iid in enumerate(self.tree.get_children()):
            bg = self.rcols[i] or (C["alt"] if i % 2 else C["surface"])
            self.tree.item(iid, tags=(f"r{i}",))
            self.tree.tag_configure(f"r{i}", background=bg, foreground=C["text"])
        stag = "sel"
        self.tree.item(sel[0], tags=(stag,))
        self.tree.tag_configure(stag, background=C["sel"], foreground="#fff")
        pre = [str(v) for v in self.raw[self.sel_row] if v is not None][:6]
        self.status.set(f"Row {self.sel_row+1}:  {' | '.join(pre)}")
        self.next_btn.config(state=tk.NORMAL)

    def _header_next(self):
        self.cur_sheet = self._sname()
        self.cur_hdr_idx = self.sel_row
        self.cur_hdr_vals = self.raw[self.cur_hdr_idx]
        self.wb.close()
        self.show_col_sel()

    # ── 3  Column selection ──────────────────────────────────────────────
    def show_col_sel(self):
        self.root.deiconify()
        self.clear()
        self.history.append("col_sel")
        bar = tk.Frame(self.root, bg=C["bar"], height=75)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text=Path(self.cur_path).name, font=F_T,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=20)
        tk.Label(bar, text=f"File {len(self.file_configs)+1}  |  Step 2 : Select columns to compare",
                 font=F_SUB, fg=C["dim"], bg=C["bar"]).pack(side=tk.RIGHT, padx=20)

        card = tk.Frame(self.root, bg=C["surface"],
                        highlightbackground=C["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=20, pady=12)
        tk.Label(card, text="Tick one or more columns:", font=F_B,
                 bg=C["surface"], fg=C["cyan"]).pack(padx=15, pady=(15, 8), anchor="w")

        wrap = tk.Frame(card, bg=C["surface"])
        wrap.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        self.col_vars = []
        ci = 0
        per_row = 4
        for val in self.cur_hdr_vals:
            if val is None:
                continue
            name = str(val).strip()
            if not name or name.lower() == "nan":
                continue
            var = tk.BooleanVar()
            cb = tk.Checkbutton(wrap, text=name, variable=var, font=F,
                                bg=C["surface"], fg=C["text"],
                                selectcolor=C["chk_sel"],
                                activebackground=C["surface"],
                                activeforeground=C["accent"],
                                padx=8, pady=4, anchor="w")
            r, c = divmod(ci, per_row)
            cb.grid(row=r, column=c, sticky="w", padx=4, pady=2)
            self.col_vars.append((name, var))
            ci += 1
        for c in range(per_row):
            wrap.columnconfigure(c, weight=1)

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=20, pady=(0, 12))
        ttk.Button(bot, text="<- Back", style="A.TButton", command=self._back).pack(side=tk.LEFT)
        ttk.Button(bot, text="Next  ->", style="A.TButton",
                   command=self._cols_next).pack(side=tk.RIGHT)

    def _cols_next(self):
        cols = [n for n, v in self.col_vars if v.get()]
        if not cols:
            messagebox.showwarning("", "Select at least one column"); return
        self.file_configs.append({"path": self.cur_path, "sheet": self.cur_sheet,
                                  "header_row": self.cur_hdr_idx, "columns": cols})
        if self.template_mode and not self.template_config:
            self.template_config = {"sheet": self.cur_sheet, "header_row": self.cur_hdr_idx, "columns": cols}
            for tfile in self.template_files[1:]:
                self.file_configs.append({"path": tfile, "sheet": self.cur_sheet,
                                         "header_row": self.cur_hdr_idx, "columns": cols})
            self.auto_generate_template_mappings()
            self.run_comparison()
        elif self.temp_files:
            self.process_next_file()
        elif messagebox.askyesno("", f"{len(self.file_configs)} file(s) added.\nAdd another file?"):
            self.pick_files()
        elif len(self.file_configs) >= 2:
            self.show_pair_selector()
        else:
            messagebox.showinfo("", "Need at least 2 files"); self.pick_files()

    # ── 4  What compares to what? ────────────────────────────────────────
    def show_pair_selector(self):
        self.root.deiconify()
        self.clear()
        self.history.append("pair_selector")
        bar = tk.Frame(self.root, bg=C["bar"], height=75)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text="What compares to what?", font=F_T,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=20)
        tk.Label(bar, text="Final step : Map columns from File #1 to each other file",
                 font=F_SUB, fg=C["dim"], bg=C["bar"]).pack(side=tk.RIGHT, padx=20)

        card = tk.Frame(self.root, bg=C["surface"],
                        highlightbackground=C["border"], highlightthickness=1)
        card.pack(fill=tk.BOTH, expand=True, padx=20, pady=12)

        f1 = self.file_configs[0]
        others = self.file_configs[1:]
        SKIP = "-- skip --"

        # scrollable area
        canvas = tk.Canvas(card, bg=C["surface"], highlightthickness=0)
        vsb = ttk.Scrollbar(card, orient=tk.VERTICAL, command=canvas.yview)
        inner = tk.Frame(canvas, bg=C["surface"])
        inner.bind("<Configure>", lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw", tags="inner")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure("inner", width=e.width))
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 4), pady=4)
        canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-e.delta // 120, "units"))
        canvas.bind_all("<Button-4>", lambda _: canvas.yview_scroll(-3, "units"))
        canvas.bind_all("<Button-5>", lambda _: canvas.yview_scroll(3, "units"))

        grid = tk.Frame(inner, bg=C["surface"])
        grid.pack(fill=tk.X, padx=8, pady=8)

        # header row
        f1_name = Path(f1["path"]).name
        tk.Label(grid, text=f"#{1}  {f1_name}", font=F_B,
                 bg=C["surface"], fg=C["accent"]).grid(row=0, column=0, sticky="w",
                                                       padx=(10, 20), pady=(0, 8))
        for k, oth in enumerate(others):
            name = Path(oth["path"]).name
            tk.Label(grid, text=f"#{k+2}  {name}", font=F_B,
                     bg=C["surface"], fg=C["orange"]).grid(row=0, column=k + 1,
                                                           sticky="w", padx=10, pady=(0, 8))

        # separator
        sep = tk.Frame(grid, bg=C["border"], height=1)
        sep.grid(row=1, column=0, columnspan=1 + len(others), sticky="ew",
                 padx=10, pady=(0, 6))

        # one row per File #1 column, with a dropdown per other file
        self.map_vars = []  # [(col1, [(file_idx, var), ...]), ...]
        for ri, col1 in enumerate(f1["columns"]):
            tk.Label(grid, text=col1, font=F, bg=C["surface"],
                     fg=C["text"]).grid(row=ri + 2, column=0, sticky="w",
                                        padx=(10, 20), pady=4)
            row_vars = []
            for k, oth in enumerate(others):
                options = [SKIP] + oth["columns"]
                var = tk.StringVar(value=SKIP)
                # auto-match by name (case-insensitive)
                for oc in oth["columns"]:
                    if oc.strip().lower() == col1.strip().lower():
                        var.set(oc); break
                cb = ttk.Combobox(grid, textvariable=var, values=options,
                                  state="readonly", width=28, font=F)
                cb.grid(row=ri + 2, column=k + 1, sticky="w", padx=10, pady=4)
                row_vars.append((k + 1, var))  # k+1 = file_configs index
            self.map_vars.append((col1, row_vars))

        for c in range(1 + len(others)):
            grid.columnconfigure(c, weight=1)

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=20, pady=(0, 12))
        ttk.Button(bot, text="<- Back", style="A.TButton", command=self._back).pack(side=tk.LEFT)
        ttk.Button(bot, text="Compare  ->", style="A.TButton",
                   command=self._pairs_next).pack(side=tk.RIGHT)

    def _pairs_next(self):
        SKIP = "-- skip --"
        # build mapping: for each other file, which col1->colN pairs
        n_others = len(self.file_configs) - 1
        self.mappings = {}  # {file_idx: [(col1, colN), ...]}
        for col1, row_vars in self.map_vars:
            for fidx, var in row_vars:
                v = var.get()
                if v != SKIP:
                    self.mappings.setdefault(fidx, []).append((col1, v))
        if not self.mappings:
            messagebox.showwarning("", "Map at least one column"); return
        self.run_comparison()

    # ── 5  Comparison results ────────────────────────────────────────────
    def run_comparison(self):
        self.root.deiconify()
        self.clear()
        bar = tk.Frame(self.root, bg=C["bar"], height=70)
        bar.pack(fill=tk.X); bar.pack_propagate(False)
        tk.Label(bar, text="Comparison Results", font=F_T,
                 fg=C["accent"], bg=C["bar"]).pack(side=tk.LEFT, padx=20, pady=12)

        loading = tk.Toplevel(self.root)
        loading.title("Loading")
        loading.geometry("300x100")
        loading.resizable(False, False)
        tk.Label(loading, text="Loading Excel files...", font=F_B,
                 bg=C["bg"], fg=C["text"]).pack(expand=True)
        self.root.update()

        dfs = []
        for cfg in self.file_configs:
            sheet = self.resolve_sheet_name(cfg["path"], cfg["sheet"])
            df = pd.read_excel(cfg["path"], header=cfg["header_row"], sheet_name=sheet)
            dfs.append(df)

        loading.destroy()

        canvas = tk.Canvas(self.root, bg=C["bg"], highlightthickness=0)
        vsb = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=canvas.yview)
        self._res_inner = tk.Frame(canvas, bg=C["bg"])
        self._res_inner.bind("<Configure>",
                             lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._res_inner, anchor="nw", tags="inner")
        canvas.bind("<Configure>", lambda e: canvas.itemconfigure("inner", width=e.width))
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)
        canvas.bind_all("<MouseWheel>",
                        lambda e: canvas.yview_scroll(-e.delta // 120, "units"))
        canvas.bind_all("<Button-4>", lambda _: canvas.yview_scroll(-3, "units"))
        canvas.bind_all("<Button-5>", lambda _: canvas.yview_scroll(3, "units"))

        f1_cfg = self.file_configs[0]
        f1_name = Path(f1_cfg["path"]).name
        summary_data = []

        for fidx, col_pairs in sorted(self.mappings.items()):
            fN_cfg = self.file_configs[fidx]
            fN_name = Path(fN_cfg["path"]).name

            col_data_1 = {c1: self.collect_col_data(dfs[0], c1, f1_cfg["header_row"])
                          for c1, _ in col_pairs}
            col_data_N = {cN: self.collect_col_data(dfs[fidx], cN, fN_cfg["header_row"])
                          for _, cN in col_pairs}

            all_vals_1 = {v for data in col_data_1.values() for v in data.values()}
            all_vals_N = {v for data in col_data_N.values() for v in data.values()}
            common     = all_vals_1 & all_vals_N
            unique_1   = all_vals_1 - all_vals_N
            unique_N   = all_vals_N - all_vals_1

            rows_only_1 = self.get_rows_with_unique_values(col_data_1, unique_1)
            rows_only_N = self.get_rows_with_unique_values(col_data_N, unique_N)

            card = tk.Frame(self._res_inner, bg=C["surface"],
                            highlightbackground=C["border"], highlightthickness=1)
            card.pack(fill=tk.X, pady=6, padx=4)

            tk.Label(card,
                     text=f"{f1_name} [{f1_cfg['sheet']}]   vs   {fN_name} [{fN_cfg['sheet']}]",
                     font=F_B, bg=C["surface"], fg=C["text"]).pack(padx=15, pady=(12, 2), anchor="w")

            mapped_str = "    ".join(f"{c1}  ->  {cN}" for c1, cN in col_pairs)
            tk.Label(card, text=mapped_str, font=F,
                     bg=C["surface"], fg=C["dim"]).pack(padx=15, pady=(0, 4), anchor="w")

            if self.template_mode:
                self.show_template_validation_card(card, fN_cfg, fN_name)

            stats = tk.Frame(card, bg=C["surface"])
            stats.pack(fill=tk.X, padx=15, pady=4)
            tk.Label(stats, text=f"Common: {len(common)}", font=F_B,
                     bg=C["surface"], fg=C["green"]).pack(side=tk.LEFT, padx=(0, 18))
            tk.Label(stats, text=f"Only in {f1_name}: {len(rows_only_1)}", font=F_B,
                     bg=C["surface"], fg=C["accent"]).pack(side=tk.LEFT, padx=(0, 18))
            tk.Label(stats, text=f"Only in {fN_name}: {len(rows_only_N)}", font=F_B,
                     bg=C["surface"], fg=C["orange"]).pack(side=tk.LEFT)

            cols_1 = [c1 for c1, _ in col_pairs]
            cols_N = [cN for _, cN in col_pairs]

            if rows_only_1:
                self.show_comparison_grid(card, f"Only in {f1_name}", cols_1,
                                          rows_only_1, C["accent"], f1_cfg["path"])
            if rows_only_N:
                self.show_comparison_grid(card, f"Only in {fN_name}", cols_N,
                                          rows_only_N, C["orange"], fN_cfg["path"])

            tk.Frame(card, bg=C["surface"], height=8).pack()
            summary_data.append((f1_name, fN_name, len(common),
                                  len(rows_only_1), len(rows_only_N), fidx + 1))

        sm = tk.Frame(self._res_inner, bg=C["hdr_bg"],
                      highlightbackground=C["border"], highlightthickness=1)
        sm.pack(fill=tk.X, pady=6, padx=4)
        tk.Label(sm, text="Summary", font=F_T, fg=C["purple"],
                 bg=C["hdr_bg"]).pack(padx=15, pady=(12, 4), anchor="w")
        for f1_n, fN_n, common_c, u1_c, uN_c, num in summary_data:
            tk.Label(sm,
                     text=f"{f1_n} vs {fN_n}:  {common_c} common,  "
                          f"{u1_c} unique to #1,  {uN_c} unique to #{num}",
                     font=F, fg=C["dim"], bg=C["hdr_bg"]).pack(padx=15, anchor="w")
        tk.Frame(sm, bg=C["hdr_bg"], height=12).pack()

        bot = tk.Frame(self.root, bg=C["bg"])
        bot.pack(fill=tk.X, padx=12, pady=8)
        ttk.Button(bot, text="New Comparison", style="A.TButton",
                   command=self._new_comparison).pack(side=tk.RIGHT)

    def collect_col_data(self, df, col_name, hdr_row):
        if col_name not in df.columns:
            return {}
        result = {}
        for idx, val in df[col_name].dropna().items():
            v = str(val).strip()
            if v and v.lower() != "nan":
                result[hdr_row + 2 + idx] = v
        return result

    def get_rows_with_unique_values(self, col_data, unique_values):
        rows = {}
        for col_name, data in col_data.items():
            for excel_row, val in data.items():
                if val in unique_values:
                    rows.setdefault(excel_row, {})[col_name] = val
        return rows

    def show_comparison_grid(self, parent, title, col_names, rows_data, color, file_path):
        fr = tk.Frame(parent, bg=C["surface"])
        fr.pack(fill=tk.X, padx=15, pady=(6, 2))

        hdr = tk.Frame(fr, bg=C["surface"])
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=f"{title}:", font=F_B, bg=C["surface"],
                 fg=color).pack(side=tk.LEFT, anchor="w")
        tk.Button(hdr, text="Open in Excel", font=F, bg=C["surface"], fg=color,
                  relief="flat", cursor="hand2", activebackground=C["alt"],
                  activeforeground=color,
                  command=lambda: self._open_excel_file(file_path)).pack(side=tk.LEFT, padx=8)

        grid_frame = tk.Frame(fr, bg=C["surface"])
        grid_frame.pack(fill=tk.X, pady=(4, 2))

        all_columns = ["Row"] + col_names
        tree_height = min(len(rows_data), 10)
        tree = ttk.Treeview(grid_frame, columns=all_columns, show="headings",
                            style="T.Treeview", height=tree_height)
        yscr = ttk.Scrollbar(grid_frame, orient=tk.VERTICAL, command=tree.yview)
        xscr = ttk.Scrollbar(grid_frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscr.set, xscrollcommand=xscr.set)

        tree.heading("Row", text="Row")
        tree.column("Row", width=60, anchor="center", stretch=False)
        for col in col_names:
            tree.heading(col, text=col)
            tree.column(col, width=200, minwidth=80)

        for i, excel_row in enumerate(sorted(rows_data)):
            values = [str(excel_row)] + [rows_data[excel_row].get(col, "") for col in col_names]
            tree.insert("", "end", values=values, tags=("alt" if i % 2 else "normal",))

        tree.tag_configure("alt", background=C["alt"])
        tree.tag_configure("normal", background=C["surface"])

        def on_double_click(event):
            item = tree.identify_row(event.y)
            col  = tree.identify_column(event.x)
            if not item or not col:
                return
            col_index = int(col.replace("#", "")) - 1
            values = tree.item(item, "values")
            if values and col_index < len(values) and values[col_index]:
                self._copy_to_clipboard(values[col_index])

        tree.bind("<Double-1>", on_double_click)

        yscr.pack(side=tk.RIGHT, fill=tk.Y)
        xscr.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(fill=tk.X)

    def _copy_to_clipboard(self, text):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        original_title = self.root.title()
        self.root.title(f"Copied: {text}")
        self.root.after(1500, lambda: self.root.title(original_title))

    def _open_excel_file(self, file_path):
        if sys.platform == "win32":
            os.startfile(file_path)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", file_path])
        else:
            subprocess.Popen(["xdg-open", file_path])

    def _new_comparison(self):
        self.file_configs = []
        self.history = []
        self.temp_files = []
        self.template_mode = False
        self.template_files = []
        self.template_config = None
        self.pick_files()

    # ── Template validation ──────────────────────────────────────────────
    def resolve_sheet_name(self, file_path, preferred_sheet):
        wb = load_workbook(file_path, data_only=True, read_only=True)
        result = preferred_sheet if preferred_sheet in wb.sheetnames else wb.sheetnames[0]
        wb.close()
        return result

    def auto_generate_template_mappings(self):
        template_cols = self.template_config["columns"]
        self.mappings = {fidx: [(c, c) for c in template_cols]
                         for fidx in range(1, len(self.file_configs))}

    def get_columns_at_row(self, file_path, sheet_name, row_index):
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        result = []
        for i, row in enumerate(ws.iter_rows(max_row=row_index + 1, max_col=40)):
            if i == row_index:
                result = [str(c.value).strip() for c in row if c.value is not None]
        wb.close()
        return result

    def find_actual_header_row(self, file_path, sheet_name, expected_columns):
        expected_lower = {c.lower() for c in expected_columns}
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        best_row, best_match = None, 0
        for row_idx, row in enumerate(ws.iter_rows(max_row=50, max_col=40)):
            row_vals = {str(c.value).strip().lower() for c in row if c.value is not None}
            match_count = len(expected_lower & row_vals)
            if match_count > best_match:
                best_match, best_row = match_count, row_idx
        wb.close()
        return best_row if best_match >= max(1, len(expected_lower) * 0.5) else None

    def validate_file_against_template(self, file_cfg):
        expected_row = self.template_config["header_row"]
        template_cols = self.template_config["columns"]
        template_sheet = self.template_config["sheet"]
        wb = load_workbook(file_cfg["path"], data_only=True, read_only=True)
        sheet_found = template_sheet in wb.sheetnames
        actual_sheet = template_sheet if sheet_found else wb.sheetnames[0]
        wb.close()
        actual_cols = self.get_columns_at_row(file_cfg["path"], actual_sheet, expected_row)
        missing = [c for c in template_cols if c not in actual_cols]
        extra = [c for c in actual_cols if c not in template_cols]
        actual_row = None
        if missing:
            found = self.find_actual_header_row(file_cfg["path"], actual_sheet, template_cols)
            actual_row = (found + 1) if found is not None else None
        return {"expected_row": expected_row + 1, "missing": missing, "extra": extra,
                "actual_row": actual_row, "sheet_found": sheet_found,
                "expected_sheet": template_sheet, "actual_sheet": actual_sheet}

    def show_template_validation_card(self, parent, file_cfg, file_name):
        v = self.validate_file_against_template(file_cfg)
        bg = C["surface2"]
        vcard = tk.Frame(parent, bg=bg, highlightbackground=C["border"], highlightthickness=1)
        vcard.pack(fill=tk.X, padx=15, pady=(2, 8))
        has_issues = not v["sheet_found"] or v["missing"]
        if not has_issues:
            tk.Label(vcard,
                     text=f"Template check \u2713  {file_name}  \u2014  sheet '{v['expected_sheet']}', row {v['expected_row']}, all columns found",
                     font=F_B, bg=bg, fg=C["green"]).pack(padx=10, pady=6, anchor="w")
            return
        tk.Label(vcard, text=f"Template mismatch \u2717  {file_name}",
                 font=F_B, bg=bg, fg=C["red"]).pack(padx=10, pady=(6, 2), anchor="w")
        if not v["sheet_found"]:
            tk.Label(vcard, text=f"  Sheet '{v['expected_sheet']}' not found  \u2014  using '{v['actual_sheet']}' instead",
                     font=F, bg=bg, fg=C["orange"]).pack(padx=10, pady=1, anchor="w")
        if v["missing"]:
            row_msg = f"row {v['expected_row']}"
            if v["actual_row"]:
                row_msg += f"  (header found at row {v['actual_row']})"
            tk.Label(vcard, text=f"  Header mismatch at {row_msg}  \u2014  missing: {',  '.join(v['missing'])}",
                     font=F, bg=bg, fg=C["orange"]).pack(padx=10, pady=1, anchor="w")
        if v["extra"]:
            tk.Label(vcard, text=f"  Unexpected columns: {',  '.join(v['extra'])}",
                     font=F, bg=bg, fg=C["dim"]).pack(padx=10, pady=(1, 6), anchor="w")

    def _back(self):
        if len(self.history) < 2:
            return
        self.history.pop()
        if self.history[-1] == "sheet_and_header":
            self.show_sheet_and_header()
        elif self.history[-1] == "col_sel":
            self.show_col_sel()
        elif self.history[-1] == "pair_selector":
            self.show_pair_selector()


if __name__ == "__main__":
    App()
# Requires: pip install openpyxl pandas
# On Linux you may also need: sudo apt install python3-tk
