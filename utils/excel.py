from openpyxl import load_workbook
import pandas as pd
from utils.text import normalize


def load_dataframe(path, header_row, sheet_name):
    df = pd.read_excel(path, header=header_row, sheet_name=sheet_name)
    df.columns = [normalize(c) for c in df.columns]
    return df


def resolve_sheet_name(file_path, preferred_sheet):
    wb = load_workbook(file_path, data_only=True, read_only=True)
    result = preferred_sheet if preferred_sheet in wb.sheetnames else wb.sheetnames[0]
    wb.close()
    return result


def get_filter_header_rows(ws):
    header_rows = set()
    for table in ws.tables.values():
        row_number = int(table.ref.split(":")[0].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        header_rows.add(row_number - 1)
    if ws.auto_filter and ws.auto_filter.ref:
        row_number = int(ws.auto_filter.ref.split(":")[0].lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        header_rows.add(row_number - 1)
    return header_rows


def get_columns_at_row(file_path, sheet_name, row_index):
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    result = []
    for i, row in enumerate(ws.iter_rows(max_row=row_index + 1, max_col=40)):
        if i == row_index:
            result = [normalize(c.value) for c in row if c.value is not None]
    wb.close()
    return result


def find_actual_header_row(file_path, sheet_name, expected_columns):
    expected_lower = {c.lower() for c in expected_columns}
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    best_row, best_match = None, 0
    for row_idx, row in enumerate(ws.iter_rows(max_row=50, max_col=40)):
        row_vals = {normalize(c.value).lower() for c in row if c.value is not None}
        match_count = len(expected_lower & row_vals)
        if match_count > best_match:
            best_match, best_row = match_count, row_idx
    wb.close()
    return best_row if best_match >= max(1, len(expected_lower) * 0.5) else None
