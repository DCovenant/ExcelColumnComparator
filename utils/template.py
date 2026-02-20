import tkinter as tk
from openpyxl import load_workbook
from utils.theme import COLORS as C, F, F_BOLD
from utils.excel import get_columns_at_row, find_actual_header_row


def validate_file_against_template(template_config, file_cfg):
    expected_row = template_config["header_row"]
    template_cols = template_config["columns"]
    template_sheet = template_config["sheet"]
    wb = load_workbook(file_cfg["path"], data_only=True, read_only=True)
    sheet_found = template_sheet in wb.sheetnames
    actual_sheet = template_sheet if sheet_found else wb.sheetnames[0]
    wb.close()
    actual_cols = get_columns_at_row(file_cfg["path"], actual_sheet, expected_row)
    missing = [c for c in template_cols if c not in actual_cols]
    extra = [c for c in actual_cols if c not in template_cols]
    actual_row = None
    if missing:
        found = find_actual_header_row(file_cfg["path"], actual_sheet, template_cols)
        actual_row = (found + 1) if found is not None else None
    return {"expected_row": expected_row + 1, "missing": missing, "extra": extra,
            "actual_row": actual_row, "sheet_found": sheet_found,
            "expected_sheet": template_sheet, "actual_sheet": actual_sheet}


def show_template_validation_card(parent, template_config, file_cfg, file_name):
    v = validate_file_against_template(template_config, file_cfg)
    bg = C["surface2"]
    vcard = tk.Frame(parent, bg=bg, highlightbackground=C["border"], highlightthickness=1)
    vcard.pack(fill=tk.X, padx=15, pady=(2, 8))
    has_issues = not v["sheet_found"] or v["missing"]
    if not has_issues:
        tk.Label(vcard,
                 text=f"Template check \u2713  {file_name}  \u2014  sheet '{v['expected_sheet']}', row {v['expected_row']}, all columns found",
                 font=F_BOLD, bg=bg, fg=C["green"]).pack(padx=10, pady=6, anchor="w")
        return
    tk.Label(vcard, text=f"Template mismatch \u2717  {file_name}",
             font=F_BOLD, bg=bg, fg=C["red"]).pack(padx=10, pady=(6, 2), anchor="w")
    if not v["sheet_found"]:
        tk.Label(vcard, text=f"  Sheet '{v['expected_sheet']}' not found  \u2014  using '{v['actual_sheet']}' instead",
                 font=F, bg=bg, fg=C["orange"]).pack(padx=10, pady=1, anchor="w")
    if v["missing"]:
        row_msg = f"row {v['expected_row']}"
        if v["actual_row"]:
            row_msg += f"  (header found at row {v['actual_row']})"
        tk.Label(vcard, text=f"  Header mismatch at {row_msg}  \u2014  missing: {',  '.join(v['missing'])}",
                 font=F, bg=bg, fg=C["orange"]).pack(padx=10, pady=1, anchor="w")
    if v["extra"]:
        tk.Label(vcard, text=f"  Unexpected columns: {',  '.join(v['extra'])}",
                 font=F, bg=bg, fg=C["dim"]).pack(padx=10, pady=(1, 6), anchor="w")
